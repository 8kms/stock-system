"""
输出多 Sheet Excel
"""
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR, SCORE_THRESHOLDS


def write_report(ranking_df, factor_df, index_detail, risk_df, kline_results,
                 importance_df=None, watchlist_codes=None):
    """
    写出完整的周报 Excel
    """
    date_str = datetime.now().strftime("%Y%m%d")
    path = OUTPUT_DIR / f"A股选股报告_{date_str}.xlsx"

    # 用 openpyxl 写多 Sheet
    wb = Workbook()

    _write_top30(wb, ranking_df)
    _write_watchlist(wb, ranking_df, watchlist_codes)
    _write_index_score(wb, index_detail)
    _write_risk_list(wb, risk_df)
    _write_factor_detail(wb, factor_df)
    _write_kline_summary(wb, kline_results, ranking_df)

    # 删除默认 Sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(path)
    print(f"报告已生成: {path}")
    return path


def _style_header(ws, row, ncols):
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def _write_top30(wb, ranking_df):
    ws = wb.create_sheet("Top30")
    top30 = ranking_df.head(30)

    headers = ["排名", "代码", "名称", "综合得分", "模型评分", "K线评分", "行业", "判定"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    for i, (_, row) in enumerate(top30.iterrows()):
        r = i + 2
        score = row.get("total_score", row.get("model_score", 50))
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=row.get("code", ""))
        ws.cell(row=r, column=3, value=row.get("name", ""))
        ws.cell(row=r, column=4, value=round(score, 1))
        ws.cell(row=r, column=5, value=round(row.get("model_score", 50), 1))
        ws.cell(row=r, column=6, value=round(row.get("kline_score", 5), 1))
        ws.cell(row=r, column=7, value=row.get("industry", ""))

        if score >= SCORE_THRESHOLDS["focus"]:
            verdict = "重点研究"
        elif score >= SCORE_THRESHOLDS["watchlist"]:
            verdict = "进入观察"
        elif score >= SCORE_THRESHOLDS["normal"]:
            verdict = "普通机会"
        else:
            verdict = "仅观察"
        ws.cell(row=r, column=8, value=verdict)

        # 条件格式
        if score >= 85:
            fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = fill

        # 对齐
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=c).border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"),
            )

    _auto_width(ws)


def _write_watchlist(wb, ranking_df, watchlist_codes=None):
    ws = wb.create_sheet("Watchlist")
    if watchlist_codes is None:
        watchlist_codes = []

    watch = ranking_df[ranking_df["code"].isin(watchlist_codes)]
    if len(watch) == 0:
        # 取分数在观察区间的
        watch = ranking_df[
            (ranking_df.get("total_score", ranking_df.get("model_score", 0)) >= SCORE_THRESHOLDS["watchlist"])
        ]

    headers = ["代码", "名称", "综合得分", "行业", "目前排名", "判定"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    for i, (_, row) in enumerate(watch.iterrows()):
        r = i + 2
        ws.cell(row=r, column=1, value=row.get("code", ""))
        ws.cell(row=r, column=2, value=row.get("name", ""))
        ws.cell(row=r, column=3, value=round(row.get("total_score", row.get("model_score", 50)), 1))
        ws.cell(row=r, column=4, value=row.get("industry", ""))
        ws.cell(row=r, column=5, value=row.get("rank", i + 1))

        score = row.get("total_score", row.get("model_score", 50))
        if score >= 85:
            verdict = "重点研究"
        elif score >= 75:
            verdict = "进入观察"
        else:
            verdict = "仅观察"
        ws.cell(row=r, column=6, value=verdict)

        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")

    _auto_width(ws)


def _write_index_score(wb, index_detail):
    ws = wb.create_sheet("IndexScore")

    # 先写总览
    if isinstance(index_detail, dict):
        ws.cell(row=1, column=1, value="市场状态总览")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"综合评分: {index_detail.get('total_score', 'N/A')} / {index_detail.get('max_score', 10)}")
        ws.cell(row=3, column=1, value=f"市场状态: {index_detail.get('state_cn', 'N/A')}")
        ws.cell(row=4, column=1, value=f"操作建议: {index_detail.get('suggestion', '')}")
        ws.cell(row=4, column=1).font = Font(color="FF0000")

        details = index_detail.get("details", [])
        if details:
            headers = list(details[0].keys())
            start_row = 6
            for col, h in enumerate(headers, 1):
                ws.cell(row=start_row, column=col, value=h)
            _style_header(ws, start_row, len(headers))

            for i, d in enumerate(details):
                r = start_row + 1 + i
                for col, h in enumerate(headers, 1):
                    ws.cell(row=r, column=col, value=d.get(h, ""))
                    ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")

    _auto_width(ws)


def _write_risk_list(wb, risk_df):
    ws = wb.create_sheet("RiskList")

    if risk_df is None or risk_df.empty:
        ws.cell(row=1, column=1, value="暂无显著风险股票")
        return

    high_risk = risk_df.sort_values("risk_score", ascending=False).head(30)
    headers = ["代码", "名称", "风险分", "风险类型", "说明"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    for i, (_, row) in enumerate(high_risk.iterrows()):
        r = i + 2
        ws.cell(row=r, column=1, value=row.get("code", ""))
        ws.cell(row=r, column=2, value=row.get("name", ""))
        ws.cell(row=r, column=3, value=round(row.get("risk_score", 0), 1))
        ws.cell(row=r, column=4, value=row.get("risk_type", ""))
        ws.cell(row=r, column=5, value=row.get("description", ""))

        if row.get("risk_score", 0) > 15:
            fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = fill

        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")

    _auto_width(ws)


def _write_factor_detail(wb, factor_df):
    ws = wb.create_sheet("FactorDetail")

    if factor_df is None or factor_df.empty:
        ws.cell(row=1, column=1, value="因子数据暂缺")
        return

    # 只写前 50 只
    subset = factor_df.head(50)
    headers = list(subset.columns)
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    for i, (_, row) in enumerate(subset.iterrows()):
        r = i + 2
        for col, h in enumerate(headers, 1):
            val = row.get(h, "")
            if isinstance(val, float):
                val = round(val, 4)
            ws.cell(row=r, column=col, value=val)
            ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")

    _auto_width(ws)


def _write_kline_summary(wb, kline_results, ranking_df):
    ws = wb.create_sheet("KlineSummary")

    headers = ["代码", "名称", "K线健康分", "是否走坏", "是否过热", "是否企稳", "波动异常", "信号"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    top_codes = ranking_df.head(50)["code"].tolist() if ranking_df is not None else []
    row_idx = 2
    for code in top_codes:
        if code in kline_results:
            kr = kline_results[code]
            ws.cell(row=row_idx, column=1, value=code)
            name = ""
            if ranking_df is not None:
                name_row = ranking_df[ranking_df["code"] == code]
                if not name_row.empty:
                    name = name_row.iloc[0].get("name", "")
            ws.cell(row=row_idx, column=2, value=name)
            ws.cell(row=row_idx, column=3, value=round(kr.get("score", 5), 1))
            ws.cell(row=row_idx, column=4, value="是" if not kr.get("is_healthy") else "否")
            ws.cell(row=row_idx, column=5, value="是" if kr.get("is_overbought") else "否")
            ws.cell(row=row_idx, column=6, value="是" if kr.get("is_stabilizing") else "否")
            ws.cell(row=row_idx, column=7, value="是" if kr.get("vol_abnormal") else "否")
            ws.cell(row=row_idx, column=8, value="; ".join(kr.get("signals", [])))

            for c in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=c).alignment = Alignment(horizontal="center")

            row_idx += 1

    _auto_width(ws)


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)
